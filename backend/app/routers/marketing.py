"""Marketing module: semantic core collection for SEO and Yandex Direct."""
from __future__ import annotations

import csv
import io
import logging
import re
import uuid
from datetime import datetime, timedelta, timezone
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.auth.deps import CurrentUser, NonViewerRequired
from app.db.session import get_db
from app.models.marketing import (
    CleaningSnapshot,
    KeywordCache,
    MarketingMinusWord,
    SemanticCluster,
    SemanticKeyword,
    SemanticMode,
    SemanticProject,
)
from app.models.project import Project
from app.models.task import Task, TaskStatus, TaskType
from app.models.user import UserRole
from app.routers.history import log_event

logger = logging.getLogger(__name__)

router = APIRouter()

CACHE_TTL_DAYS = 30


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _check_project_access(project_id: uuid.UUID, current_user, db: Session) -> Project:
    project = db.get(Project, project_id)
    if not project or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role == UserRole.SPECIALIST and project.specialist_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    return project


def _get_sem_project(sem_id: uuid.UUID, project_id: uuid.UUID, db: Session) -> SemanticProject:
    sp = db.get(SemanticProject, sem_id)
    if not sp or sp.project_id != project_id:
        raise HTTPException(status_code=404, detail="Semantic project not found")
    return sp


def _classify_kw_type(exact: int | None) -> str | None:
    if exact is None:
        return None
    if exact >= 1000:
        return "ВЧ"
    if exact >= 100:
        return "СЧ"
    return "НЧ"


# ─── Schemas ──────────────────────────────────────────────────────────────────

class SemanticProjectCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    mode: SemanticMode
    region: str | None = Field(None, max_length=100)
    region_id: int | None = None
    is_seasonal: bool = False


class SemanticProjectResponse(BaseModel):
    id: uuid.UUID
    project_id: uuid.UUID
    name: str
    mode: str
    region: str | None
    region_id: int | None
    is_seasonal: bool
    needs_brand_check: bool
    pipeline_step: int
    created_at: datetime
    updated_at: datetime

    model_config = {"from_attributes": True}


class CollectMasksRequest(BaseModel):
    masks: list[str] = Field(..., min_length=1, max_length=50)

    @classmethod
    def __get_validators__(cls):
        yield cls._validate

    @classmethod
    def _validate(cls, v):
        return v


class SemanticKeywordResponse(BaseModel):
    id: uuid.UUID
    phrase: str
    frequency_base: int | None
    frequency_phrase: int | None
    frequency_exact: int | None
    frequency_order: int | None
    kw_type: str | None
    intent: str | None
    source: str
    is_mask: bool
    mask_selected: bool
    is_branded: bool
    is_competitor: bool
    is_seasonal: bool
    geo_dependent: bool
    is_excluded: bool
    cluster_name: str | None
    created_at: datetime

    model_config = {"from_attributes": True}


class MaskSelectionUpdate(BaseModel):
    mask_selected: bool


class KeywordsListResponse(BaseModel):
    items: list[SemanticKeywordResponse]
    total: int
    page: int
    per_page: int


# ─── Semantic Project CRUD ─────────────────────────────────────────────────────

@router.post(
    "/projects/{project_id}/marketing/semantic",
    response_model=SemanticProjectResponse,
    status_code=201,
)
def create_or_get_semantic_project(
    project_id: uuid.UUID,
    body: SemanticProjectCreate,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Create a semantic project for a given mode. One per mode per project (UPSERT)."""
    _check_project_access(project_id, current_user, db)

    existing = db.scalar(
        select(SemanticProject).where(
            SemanticProject.project_id == project_id,
            SemanticProject.mode == body.mode,
        )
    )
    if existing:
        # Update name/region if provided
        existing.name = body.name
        existing.region = body.region
        existing.region_id = body.region_id
        existing.is_seasonal = body.is_seasonal
        db.commit()
        db.refresh(existing)
        return existing

    sp = SemanticProject(
        project_id=project_id,
        name=body.name,
        mode=body.mode,
        region=body.region,
        region_id=body.region_id,
        is_seasonal=body.is_seasonal,
        pipeline_step=0,
    )
    db.add(sp)
    db.commit()
    db.refresh(sp)
    log_event(project_id, current_user, "semantic_created", f"Создано семантическое ядро «{sp.name}» ({body.mode})", db)
    return sp


@router.get(
    "/projects/{project_id}/marketing/semantic",
    response_model=list[SemanticProjectResponse],
)
def list_semantic_projects(
    project_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    _check_project_access(project_id, current_user, db)
    return db.scalars(
        select(SemanticProject)
        .where(SemanticProject.project_id == project_id)
        .order_by(SemanticProject.created_at)
    ).all()


@router.get(
    "/projects/{project_id}/marketing/semantic/{sem_id}",
    response_model=SemanticProjectResponse,
)
def get_semantic_project(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    _check_project_access(project_id, current_user, db)
    return _get_sem_project(sem_id, project_id, db)


@router.delete("/projects/{project_id}/marketing/semantic/{sem_id}", status_code=204)
def delete_semantic_project(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    _check_project_access(project_id, current_user, db)
    sp = _get_sem_project(sem_id, project_id, db)
    # Delete child records first (no CASCADE on FK)
    from sqlalchemy import delete as sa_delete
    db.execute(sa_delete(MarketingMinusWord).where(MarketingMinusWord.semantic_project_id == sem_id))
    db.execute(sa_delete(SemanticCluster).where(SemanticCluster.semantic_project_id == sem_id))
    db.execute(sa_delete(CleaningSnapshot).where(CleaningSnapshot.semantic_project_id == sem_id))
    db.execute(sa_delete(SemanticKeyword).where(SemanticKeyword.semantic_project_id == sem_id))
    sp_name = sp.name
    db.delete(sp)
    db.commit()
    log_event(project_id, current_user, "semantic_deleted", f"Удалено семантическое ядро «{sp_name}»", db)


# ─── Mask Collection ───────────────────────────────────────────────────────────

@router.post(
    "/projects/{project_id}/marketing/semantic/{sem_id}/collect-masks",
    response_model=list[SemanticKeywordResponse],
)
async def collect_masks(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    body: CollectMasksRequest,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Collect all 4 frequency types for mask phrases via Wordstat.

    Uses cache (30-day TTL) to avoid redundant API calls.
    Saves results as SemanticKeyword records with is_mask=True.
    """
    _check_project_access(project_id, current_user, db)
    sp = _get_sem_project(sem_id, project_id, db)

    # Normalize and deduplicate
    raw_masks = [m.strip().lower() for m in body.masks if m.strip()]
    masks = list(dict.fromkeys(raw_masks))  # preserve order, remove dupes
    if not masks:
        raise HTTPException(status_code=422, detail="No valid masks provided")

    from app.services.wordstat import get_wordstat_client

    client = get_wordstat_client(db)

    # ── Cache lookup ──────────────────────────────────────────────────────────
    cutoff = datetime.now(tz=timezone.utc) - timedelta(days=CACHE_TTL_DAYS)
    cached_rows = db.scalars(
        select(KeywordCache).where(
            KeywordCache.phrase.in_(masks),
            KeywordCache.region_id == sp.region_id,
            KeywordCache.cached_at > cutoff,
        )
    ).all()
    # Skip all-zero cache entries — they likely came from failed API calls
    cached_map: dict[str, KeywordCache] = {
        row.phrase: row
        for row in cached_rows
        if (row.frequency_base or 0) + (row.frequency_phrase or 0)
        + (row.frequency_exact or 0) + (row.frequency_order or 0) > 0
    }
    uncached = [m for m in masks if m not in cached_map]

    # ── Fetch uncached from Wordstat ──────────────────────────────────────────
    fresh: dict[str, dict] = {}
    if uncached:
        if not client:
            raise HTTPException(
                status_code=503,
                detail="Wordstat API key or folder ID not configured. Add them in Settings → API keys.",
            )
        regions = [sp.region_id] if sp.region_id else None
        try:
            fresh = await client.get_all_frequencies(uncached, regions=regions)
        except Exception as exc:
            logger.exception("Wordstat error: %s", exc)
            raise HTTPException(status_code=502, detail=f"Wordstat API error: {exc}")

        # Save to cache (skip all-zero results — likely API errors)
        now = datetime.now(tz=timezone.utc)
        for phrase, freqs in fresh.items():
            total = freqs["base"] + freqs["phrase_freq"] + freqs["exact"] + freqs["order"]
            if total == 0:
                continue  # Don't cache zeros from failed/empty API responses
            existing_cache = db.scalar(
                select(KeywordCache).where(
                    KeywordCache.phrase == phrase,
                    KeywordCache.region_id == sp.region_id,
                )
            )
            if existing_cache:
                existing_cache.frequency_base = freqs["base"]
                existing_cache.frequency_phrase = freqs["phrase_freq"]
                existing_cache.frequency_exact = freqs["exact"]
                existing_cache.frequency_order = freqs["order"]
                existing_cache.cached_at = now
            else:
                db.add(KeywordCache(
                    phrase=phrase,
                    region_id=sp.region_id,
                    frequency_base=freqs["base"],
                    frequency_phrase=freqs["phrase_freq"],
                    frequency_exact=freqs["exact"],
                    frequency_order=freqs["order"],
                    cached_at=now,
                ))

    # ── Build combined frequency map ──────────────────────────────────────────
    def _freqs(phrase: str) -> dict:
        if phrase in fresh:
            return fresh[phrase]
        if phrase in cached_map:
            c = cached_map[phrase]
            return {
                "base": c.frequency_base or 0,
                "phrase_freq": c.frequency_phrase or 0,
                "exact": c.frequency_exact or 0,
                "order": c.frequency_order or 0,
            }
        return {"base": 0, "phrase_freq": 0, "exact": 0, "order": 0}

    # ── Delete old mask records for this project and re-create ────────────────
    old_masks = db.scalars(
        select(SemanticKeyword).where(
            SemanticKeyword.semantic_project_id == sem_id,
            SemanticKeyword.is_mask.is_(True),
        )
    ).all()
    # Keep records for phrases still in the new list, remove others
    existing_mask_phrases = {kw.phrase: kw for kw in old_masks}
    phrases_to_remove = set(existing_mask_phrases) - set(masks)
    for phrase in phrases_to_remove:
        db.delete(existing_mask_phrases[phrase])

    result_keywords: list[SemanticKeyword] = []
    for phrase in masks:
        f = _freqs(phrase)
        kw_type = _classify_kw_type(f["exact"])

        if phrase in existing_mask_phrases:
            kw = existing_mask_phrases[phrase]
            kw.frequency_base = f["base"]
            kw.frequency_phrase = f["phrase_freq"]
            kw.frequency_exact = f["exact"]
            kw.frequency_order = f["order"]
            kw.kw_type = kw_type
        else:
            kw = SemanticKeyword(
                semantic_project_id=sem_id,
                phrase=phrase,
                frequency_base=f["base"],
                frequency_phrase=f["phrase_freq"],
                frequency_exact=f["exact"],
                frequency_order=f["order"],
                kw_type=kw_type,
                source="wordstat",
                is_mask=True,
                mask_selected=True,
            )
            db.add(kw)
        result_keywords.append(kw)

    # Advance pipeline step if still at 0
    if sp.pipeline_step < 1:
        sp.pipeline_step = 1

    db.commit()
    for kw in result_keywords:
        db.refresh(kw)
    log_event(project_id, current_user, "semantic_masks_collected", f"Собрано {len(result_keywords)} масок", db)
    return result_keywords


# ─── Mask Selection Toggle ────────────────────────────────────────────────────

@router.patch(
    "/projects/{project_id}/marketing/semantic/{sem_id}/keywords/{kw_id}/mask-selection",
    response_model=SemanticKeywordResponse,
)
def update_mask_selection(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    kw_id: uuid.UUID,
    body: MaskSelectionUpdate,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)
    kw = db.get(SemanticKeyword, kw_id)
    if not kw or kw.semantic_project_id != sem_id:
        raise HTTPException(status_code=404, detail="Keyword not found")
    kw.mask_selected = body.mask_selected
    db.commit()
    db.refresh(kw)
    return kw


# ─── Keywords List ────────────────────────────────────────────────────────────

@router.get(
    "/projects/{project_id}/marketing/semantic/{sem_id}/keywords",
    response_model=KeywordsListResponse,
)
def list_keywords(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    page: int = Query(1, ge=1),
    per_page: int = Query(100, ge=1, le=500),
    kw_type: str | None = Query(None),
    intent: str | None = Query(None),
    source: str | None = Query(None),
    only_masks: bool = Query(False),
    search: str | None = Query(None, max_length=200),
):
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)

    q = select(SemanticKeyword).where(
        SemanticKeyword.semantic_project_id == sem_id,
        SemanticKeyword.is_excluded.is_(False),
    )
    if kw_type:
        q = q.where(SemanticKeyword.kw_type == kw_type)
    if intent:
        q = q.where(SemanticKeyword.intent == intent)
    if source:
        q = q.where(SemanticKeyword.source == source)
    if only_masks:
        q = q.where(SemanticKeyword.is_mask.is_(True))
    if search:
        q = q.where(SemanticKeyword.phrase.ilike(f"%{search}%"))

    from sqlalchemy import func
    total = db.scalar(select(func.count()).select_from(q.subquery())) or 0

    items = db.scalars(
        q.order_by(SemanticKeyword.frequency_exact.desc().nullslast())
        .offset((page - 1) * per_page)
        .limit(per_page)
    ).all()

    return {"items": items, "total": total, "page": page, "per_page": per_page}


# ─── Autopilot ────────────────────────────────────────────────────────────────

class AutopilotRequest(BaseModel):
    min_freq_exact: int = Field(0, ge=0, description="Минимальная точная частотность")


class TaskResponse(BaseModel):
    task_id: str
    status: str


@router.post(
    "/projects/{project_id}/marketing/semantic/{sem_id}/autopilot",
    response_model=TaskResponse,
    status_code=202,
)
def start_autopilot(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    body: AutopilotRequest,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Launch full autopilot: brief -> masks -> expand -> clean -> cluster."""
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)

    # Check brief exists
    from app.models.brief import Brief
    brief = db.scalar(select(Brief).where(Brief.project_id == project_id))
    if not brief or not (brief.niche or brief.products):
        raise HTTPException(status_code=422, detail="Заполните бриф (ниша или продукты) перед запуском автопилота.")

    # Check no running autopilot/expand/cluster
    running = db.scalar(
        select(Task).where(
            Task.project_id == project_id,
            Task.type.in_([TaskType.SEMANTIC_AUTOPILOT, TaskType.SEMANTIC_EXPAND, TaskType.SEMANTIC_CLUSTER]),
            Task.status.in_([TaskStatus.PENDING, TaskStatus.RUNNING]),
        )
    )
    if running:
        raise HTTPException(status_code=409, detail="Задача уже запущена. Дождитесь завершения.")

    now = datetime.now(timezone.utc)
    task = Task(
        project_id=project_id,
        type=TaskType.SEMANTIC_AUTOPILOT,
        status=TaskStatus.PENDING,
        progress=0,
        created_at=now,
    )
    db.add(task)
    db.commit()
    db.refresh(task)

    from app.tasks.marketing import task_semantic_autopilot
    task_semantic_autopilot.delay(
        task_id=str(task.id),
        sem_project_id=str(sem_id),
        project_id=str(project_id),
        min_freq_exact=body.min_freq_exact,
    )

    log_event(project_id, current_user, "semantic_autopilot", "Запущен автопилот семантического ядра", db)
    return {"task_id": str(task.id), "status": "pending"}


# ─── Semantic Expansion ───────────────────────────────────────────────────────

class ExpandRequest(BaseModel):
    min_freq_exact: int = Field(0, ge=0, description="Минимальная точная частотность (0 = не фильтровать)")
    use_brief: bool = Field(True, description="Использовать контекст бриф при генерации")


@router.post(
    "/projects/{project_id}/marketing/semantic/{sem_id}/expand",
    response_model=TaskResponse,
    status_code=202,
)
def start_expand(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    body: ExpandRequest,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Start async semantic expansion via Claude + Wordstat. Returns task_id to poll."""
    _check_project_access(project_id, current_user, db)
    sp = _get_sem_project(sem_id, project_id, db)

    if sp.pipeline_step < 1:
        raise HTTPException(
            status_code=422,
            detail="Сначала соберите статистику масок (шаг 2)",
        )

    # Check at least one mask is selected
    selected_count = db.scalar(
        select(SemanticKeyword).where(
            SemanticKeyword.semantic_project_id == sem_id,
            SemanticKeyword.is_mask.is_(True),
            SemanticKeyword.mask_selected.is_(True),
        ).with_only_columns(SemanticKeyword.id)
    )
    if not selected_count:
        raise HTTPException(status_code=422, detail="Нет выбранных масок")

    # Prevent concurrent expand tasks for the same semantic project
    running_task = db.scalar(
        select(Task).where(
            Task.project_id == project_id,
            Task.type == TaskType.SEMANTIC_EXPAND,
            Task.status.in_([TaskStatus.PENDING, TaskStatus.RUNNING]),
        )
    )
    if running_task:
        raise HTTPException(
            status_code=409,
            detail="Расширение уже запущено. Дождитесь завершения.",
        )

    now = datetime.now(timezone.utc)
    task = Task(
        project_id=project_id,
        type=TaskType.SEMANTIC_EXPAND,
        status=TaskStatus.PENDING,
        progress=0,
        created_at=now,
    )
    db.add(task)
    db.commit()
    db.refresh(task)

    from app.tasks.marketing import task_semantic_expand
    task_semantic_expand.delay(
        task_id=str(task.id),
        sem_project_id=str(sem_id),
        project_id=str(project_id),
        min_freq_exact=body.min_freq_exact,
        use_brief=body.use_brief,
    )

    return {"task_id": str(task.id), "status": "pending"}


# ─── Cleaning ─────────────────────────────────────────────────────────────────

class KeywordUpdate(BaseModel):
    is_excluded: bool | None = None
    is_branded: bool | None = None
    is_competitor: bool | None = None
    is_seasonal: bool | None = None
    geo_dependent: bool | None = None
    intent: str | None = None


class AutoCleanStats(BaseModel):
    excluded_zero_freq: int
    excluded_long_tail: int
    excluded_minus_words: int
    total_excluded: int
    total_kept: int
    snapshot_id: str


class MinusWordCreate(BaseModel):
    word: str = Field(..., min_length=1, max_length=255)
    note: str | None = Field(None, max_length=255)


class MinusWordResponse(BaseModel):
    id: uuid.UUID
    word: str
    note: str | None
    added_at: datetime

    model_config = {"from_attributes": True}


def _save_snapshot(sem_id: uuid.UUID, db: Session, description: str = "авто-очистка") -> CleaningSnapshot:
    """Snapshot current exclusion state of all non-mask keywords."""
    keywords = db.scalars(
        select(SemanticKeyword).where(
            SemanticKeyword.semantic_project_id == sem_id,
        )
    ).all()
    snapshot_data = [
        {
            "id": str(kw.id),
            "phrase": kw.phrase,
            "is_excluded": kw.is_excluded,
            "is_branded": kw.is_branded,
            "is_competitor": kw.is_competitor,
            "is_seasonal": kw.is_seasonal,
            "geo_dependent": kw.geo_dependent,
            "intent": kw.intent,
        }
        for kw in keywords
    ]
    snap = CleaningSnapshot(
        semantic_project_id=sem_id,
        snapshot=snapshot_data,
        description=description,
        created_at=datetime.now(timezone.utc),
    )
    db.add(snap)
    return snap


@router.patch(
    "/projects/{project_id}/marketing/semantic/{sem_id}/keywords/{kw_id}",
    response_model=SemanticKeywordResponse,
)
def update_keyword(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    kw_id: uuid.UUID,
    body: KeywordUpdate,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)
    kw = db.get(SemanticKeyword, kw_id)
    if not kw or kw.semantic_project_id != sem_id:
        raise HTTPException(status_code=404, detail="Keyword not found")
    if body.is_excluded is not None:
        kw.is_excluded = body.is_excluded
        kw.excluded_at = datetime.now(timezone.utc) if body.is_excluded else None
    if body.is_branded is not None:
        kw.is_branded = body.is_branded
    if body.is_competitor is not None:
        kw.is_competitor = body.is_competitor
    if body.is_seasonal is not None:
        kw.is_seasonal = body.is_seasonal
    if body.geo_dependent is not None:
        kw.geo_dependent = body.geo_dependent
    if body.intent is not None:
        kw.intent = body.intent
    db.commit()
    db.refresh(kw)
    return kw


@router.post(
    "/projects/{project_id}/marketing/semantic/{sem_id}/auto-clean",
    response_model=AutoCleanStats,
)
def auto_clean(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Auto-exclude: zero exact frequency, >7 word phrases, minus-word matches.
    Saves a snapshot before applying so results can be reviewed.
    """
    _check_project_access(project_id, current_user, db)
    sp = _get_sem_project(sem_id, project_id, db)

    if sp.pipeline_step < 2:
        raise HTTPException(status_code=422, detail="Сначала выполните расширение (шаг 3)")

    # Load active minus words
    minus_words = [
        mw.word.lower()
        for mw in db.scalars(
            select(MarketingMinusWord).where(MarketingMinusWord.semantic_project_id == sem_id)
        ).all()
    ]

    # Load all non-mask keywords that aren't already excluded
    keywords = db.scalars(
        select(SemanticKeyword).where(
            SemanticKeyword.semantic_project_id == sem_id,
            SemanticKeyword.is_mask.is_(False),
            SemanticKeyword.is_excluded.is_(False),
        )
    ).all()

    # Snapshot before cleaning
    snap = _save_snapshot(sem_id, db, description="авто-очистка")

    excluded_zero = 0
    excluded_long = 0
    excluded_minus = 0
    now = datetime.now(timezone.utc)

    for kw in keywords:
        reason: str | None = None

        # Zero exact frequency (skip if project is seasonal — might be seasonal keyword)
        exact = kw.frequency_exact or 0
        if exact == 0 and not sp.is_seasonal:
            reason = "zero"

        # Very long tail (>7 words)
        if reason is None and len(kw.phrase.split()) > 7:
            reason = "long"

        # Contains minus word
        if reason is None and minus_words:
            phrase_lower = kw.phrase.lower()
            for mw in minus_words:
                if mw in phrase_lower.split():
                    reason = "minus"
                    break

        if reason:
            kw.is_excluded = True
            kw.excluded_at = now
            if reason == "zero":
                excluded_zero += 1
            elif reason == "long":
                excluded_long += 1
            else:
                excluded_minus += 1

    db.commit()
    db.refresh(snap)

    total_excluded = excluded_zero + excluded_long + excluded_minus
    total_kept = len(keywords) - total_excluded

    log_event(project_id, current_user, "semantic_clean", f"Авто-очистка: исключено {total_excluded}, оставлено {total_kept}", db)

    return {
        "excluded_zero_freq": excluded_zero,
        "excluded_long_tail": excluded_long,
        "excluded_minus_words": excluded_minus,
        "total_excluded": total_excluded,
        "total_kept": total_kept,
        "snapshot_id": str(snap.id),
    }


@router.post(
    "/projects/{project_id}/marketing/semantic/{sem_id}/cleaning/complete",
    response_model=SemanticProjectResponse,
)
def complete_cleaning(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Mark cleaning as done — advance pipeline_step to 3."""
    _check_project_access(project_id, current_user, db)
    sp = _get_sem_project(sem_id, project_id, db)
    if sp.pipeline_step < 2:
        raise HTTPException(status_code=422, detail="Сначала выполните расширение")
    _save_snapshot(sem_id, db, description="завершение очистки")
    sp.pipeline_step = max(sp.pipeline_step, 3)
    db.commit()
    db.refresh(sp)
    return sp


# ─── Minus Words ──────────────────────────────────────────────────────────────

@router.get(
    "/projects/{project_id}/marketing/semantic/{sem_id}/minus-words",
    response_model=list[MinusWordResponse],
)
def list_minus_words(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)
    return db.scalars(
        select(MarketingMinusWord)
        .where(MarketingMinusWord.semantic_project_id == sem_id)
        .order_by(MarketingMinusWord.added_at)
    ).all()


@router.post(
    "/projects/{project_id}/marketing/semantic/{sem_id}/minus-words",
    response_model=list[MinusWordResponse],
    status_code=201,
)
def add_minus_words(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    body: list[MinusWordCreate],
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Add one or more minus words (bulk)."""
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)

    if not body:
        raise HTTPException(status_code=422, detail="Нет слов")

    now = datetime.now(timezone.utc)
    existing_words = {
        mw.word.lower()
        for mw in db.scalars(
            select(MarketingMinusWord).where(MarketingMinusWord.semantic_project_id == sem_id)
        ).all()
    }

    added: list[MarketingMinusWord] = []
    for item in body:
        word = item.word.strip().lower()
        if not word or word in existing_words:
            continue
        mw = MarketingMinusWord(
            semantic_project_id=sem_id,
            word=word,
            note=item.note,
            added_at=now,
        )
        db.add(mw)
        added.append(mw)
        existing_words.add(word)

    db.commit()
    for mw in added:
        db.refresh(mw)
    return added


@router.delete(
    "/projects/{project_id}/marketing/semantic/{sem_id}/minus-words/{word_id}",
    status_code=204,
)
def delete_minus_word(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    word_id: uuid.UUID,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)
    mw = db.get(MarketingMinusWord, word_id)
    if not mw or mw.semantic_project_id != sem_id:
        raise HTTPException(status_code=404, detail="Minus word not found")
    db.delete(mw)
    db.commit()


# ─── Cleaning Snapshots ───────────────────────────────────────────────────────

class SnapshotResponse(BaseModel):
    id: uuid.UUID
    description: str
    created_at: datetime
    keyword_count: int

    model_config = {"from_attributes": True}


@router.get(
    "/projects/{project_id}/marketing/semantic/{sem_id}/cleaning/snapshots",
    response_model=list[SnapshotResponse],
)
def list_snapshots(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)
    snaps = db.scalars(
        select(CleaningSnapshot)
        .where(CleaningSnapshot.semantic_project_id == sem_id)
        .order_by(CleaningSnapshot.created_at.desc())
        .limit(20)
    ).all()
    return [
        {
            "id": s.id,
            "description": s.description,
            "created_at": s.created_at,
            "keyword_count": len(s.snapshot) if s.snapshot else 0,
        }
        for s in snaps
    ]


@router.post(
    "/projects/{project_id}/marketing/semantic/{sem_id}/cleaning/restore/{snapshot_id}",
    response_model=dict,
)
def restore_snapshot(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    snapshot_id: uuid.UUID,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Restore keyword exclusion/flag state from a snapshot."""
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)

    snap = db.get(CleaningSnapshot, snapshot_id)
    if not snap or snap.semantic_project_id != sem_id:
        raise HTTPException(status_code=404, detail="Snapshot not found")

    restored = 0
    for entry in snap.snapshot or []:
        kw_id_str = entry.get("id")
        if not kw_id_str:
            continue
        try:
            kw = db.get(SemanticKeyword, uuid.UUID(kw_id_str))
        except (ValueError, Exception):
            continue
        if not kw or kw.semantic_project_id != sem_id:
            continue
        kw.is_excluded = entry.get("is_excluded", False)
        kw.excluded_at = datetime.now(timezone.utc) if kw.is_excluded else None
        kw.is_branded = entry.get("is_branded", False)
        kw.is_competitor = entry.get("is_competitor", False)
        kw.is_seasonal = entry.get("is_seasonal", False)
        kw.geo_dependent = entry.get("geo_dependent", False)
        kw.intent = entry.get("intent")
        restored += 1

    db.commit()
    return {"restored": restored}


# ─── Clustering ───────────────────────────────────────────────────────────────

class ClusterResponse(BaseModel):
    id: uuid.UUID
    name: str
    intent: str | None
    priority: str | None
    campaign_type: str | None
    suggested_title: str | None
    suggested_description: str | None
    keyword_count: int
    created_at: datetime
    updated_at: datetime

    model_config = {"from_attributes": True}


class ClusterUpdate(BaseModel):
    name: str | None = Field(None, min_length=1, max_length=255)
    intent: str | None = None
    priority: str | None = None
    campaign_type: str | None = None
    suggested_title: str | None = None
    suggested_description: str | None = None


def _cluster_response(cluster: SemanticCluster, db: Session) -> dict:
    from sqlalchemy import func
    count = db.scalar(
        select(func.count()).select_from(
            select(SemanticKeyword.id).where(
                SemanticKeyword.semantic_project_id == cluster.semantic_project_id,
                SemanticKeyword.cluster_name == cluster.name,
                SemanticKeyword.is_excluded.is_(False),
            ).subquery()
        )
    ) or 0
    return {
        "id": cluster.id,
        "name": cluster.name,
        "intent": cluster.intent,
        "priority": cluster.priority,
        "campaign_type": cluster.campaign_type,
        "suggested_title": cluster.suggested_title,
        "suggested_description": cluster.suggested_description,
        "keyword_count": count,
        "created_at": cluster.created_at,
        "updated_at": cluster.updated_at,
    }


@router.post(
    "/projects/{project_id}/marketing/semantic/{sem_id}/cluster",
    response_model=TaskResponse,
    status_code=202,
)
def start_cluster(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Start async clustering via Claude. Returns task_id to poll."""
    _check_project_access(project_id, current_user, db)
    sp = _get_sem_project(sem_id, project_id, db)

    if sp.pipeline_step < 3:
        raise HTTPException(status_code=422, detail="Сначала завершите очистку (шаг 4)")

    # Prevent concurrent cluster tasks
    running_task = db.scalar(
        select(Task).where(
            Task.project_id == project_id,
            Task.type == TaskType.SEMANTIC_CLUSTER,
            Task.status.in_([TaskStatus.PENDING, TaskStatus.RUNNING]),
        )
    )
    if running_task:
        raise HTTPException(
            status_code=409,
            detail="Кластеризация уже запущена. Дождитесь завершения.",
        )

    now = datetime.now(timezone.utc)
    task = Task(
        project_id=project_id,
        type=TaskType.SEMANTIC_CLUSTER,
        status=TaskStatus.PENDING,
        progress=0,
        created_at=now,
    )
    db.add(task)
    db.commit()
    db.refresh(task)

    from app.tasks.marketing import task_semantic_cluster
    task_semantic_cluster.delay(
        task_id=str(task.id),
        sem_project_id=str(sem_id),
        project_id=str(project_id),
    )

    log_event(project_id, current_user, "semantic_cluster", "Запущена кластеризация семантического ядра", db)
    return {"task_id": str(task.id), "status": "pending"}


@router.get(
    "/projects/{project_id}/marketing/semantic/{sem_id}/clusters",
    response_model=list[ClusterResponse],
)
def list_clusters(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)
    clusters = db.scalars(
        select(SemanticCluster)
        .where(SemanticCluster.semantic_project_id == sem_id)
        .order_by(SemanticCluster.created_at)
    ).all()
    return [_cluster_response(c, db) for c in clusters]


@router.patch(
    "/projects/{project_id}/marketing/semantic/{sem_id}/clusters/{cluster_id}",
    response_model=ClusterResponse,
)
def update_cluster(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    cluster_id: uuid.UUID,
    body: ClusterUpdate,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)
    cluster = db.get(SemanticCluster, cluster_id)
    if not cluster or cluster.semantic_project_id != sem_id:
        raise HTTPException(status_code=404, detail="Cluster not found")

    old_name = cluster.name
    if body.name is not None:
        cluster.name = body.name
        # Update cluster_name on keywords too
        for kw in db.scalars(
            select(SemanticKeyword).where(
                SemanticKeyword.semantic_project_id == sem_id,
                SemanticKeyword.cluster_name == old_name,
            )
        ).all():
            kw.cluster_name = body.name
    if body.intent is not None:
        cluster.intent = body.intent
    if body.priority is not None:
        cluster.priority = body.priority
    if body.campaign_type is not None:
        cluster.campaign_type = body.campaign_type
    if body.suggested_title is not None:
        cluster.suggested_title = body.suggested_title
    if body.suggested_description is not None:
        cluster.suggested_description = body.suggested_description

    db.commit()
    db.refresh(cluster)
    return _cluster_response(cluster, db)


@router.delete(
    "/projects/{project_id}/marketing/semantic/{sem_id}/clusters/{cluster_id}",
    status_code=204,
)
def delete_cluster(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    cluster_id: uuid.UUID,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)
    cluster = db.get(SemanticCluster, cluster_id)
    if not cluster or cluster.semantic_project_id != sem_id:
        raise HTTPException(status_code=404, detail="Cluster not found")
    # Clear cluster_name on keywords
    for kw in db.scalars(
        select(SemanticKeyword).where(
            SemanticKeyword.semantic_project_id == sem_id,
            SemanticKeyword.cluster_name == cluster.name,
        )
    ).all():
        kw.cluster_name = None
    db.delete(cluster)
    db.commit()


# ─── Export ───────────────────────────────────────────────────────────────────

def _safe_filename(name: str) -> str:
    safe = re.sub(r"[^\w\s\-]", "", name, flags=re.UNICODE).strip()
    return re.sub(r"\s+", "_", safe)[:50] or "semantic"


def _load_export_data(sem_id: uuid.UUID, db: Session) -> tuple[list, list]:
    """Return (keywords, clusters) for export, keywords sorted by cluster then freq."""
    keywords = db.scalars(
        select(SemanticKeyword)
        .where(
            SemanticKeyword.semantic_project_id == sem_id,
            SemanticKeyword.is_excluded.is_(False),
            SemanticKeyword.is_mask.is_(False),
        )
        .order_by(
            SemanticKeyword.cluster_name.asc().nullslast(),
            SemanticKeyword.frequency_exact.desc().nullslast(),
        )
    ).all()
    clusters = db.scalars(
        select(SemanticCluster)
        .where(SemanticCluster.semantic_project_id == sem_id)
        .order_by(SemanticCluster.name)
    ).all()
    return list(keywords), list(clusters)


def _build_xlsx(sp: SemanticProject, keywords: list, clusters: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Семантическое ядро ───────────────────────────────────────────
    ws = wb.active
    ws.title = "Семантическое ядро"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    alt_fill = PatternFill("solid", fgColor="EBF3FB")

    cols = ["Фраза", "Кластер", "Тип", "Интент", "WS", '«WS»', '"!WS"', "[WS]",
            "Бренд", "Конкурент", "Сезонный", "Гео"]
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for ri, kw in enumerate(keywords, 2):
        row_fill = alt_fill if ri % 2 == 0 else None
        values = [
            kw.phrase,
            kw.cluster_name or "—",
            kw.kw_type or "—",
            kw.intent or "—",
            kw.frequency_base,
            kw.frequency_phrase,
            kw.frequency_exact,
            kw.frequency_order,
            "Да" if kw.is_branded else "",
            "Да" if kw.is_competitor else "",
            "Да" if kw.is_seasonal else "",
            "Да" if kw.geo_dependent else "",
        ]
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if row_fill:
                cell.fill = row_fill

    # Column widths
    widths = [45, 30, 6, 16, 10, 10, 10, 10, 7, 10, 9, 7]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # ── Sheet 2: Кластеры ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Кластеры")
    cluster_hdrs = ["Кластер", "Интент", "Приоритет", "Кол-во ключей"]
    if str(sp.mode.value if hasattr(sp.mode, 'value') else sp.mode) == "direct":
        cluster_hdrs += ["Тип кампании", "Заголовок объявления"]
    for ci, col in enumerate(cluster_hdrs, 1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font

    # Count per cluster
    cluster_count: dict[str, int] = {}
    for kw in keywords:
        key = kw.cluster_name or "—"
        cluster_count[key] = cluster_count.get(key, 0) + 1

    for ri, c in enumerate(clusters, 2):
        row = [c.name, c.intent or "", c.priority or "", cluster_count.get(c.name, 0)]
        if str(sp.mode.value if hasattr(sp.mode, 'value') else sp.mode) == "direct":
            row += [c.campaign_type or "", c.suggested_title or ""]
        for ci, v in enumerate(row, 1):
            ws2.cell(row=ri, column=ci, value=v)

    for ci, w in enumerate([35, 18, 12, 12, 14, 40], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3 (Direct mode): Кампании ──────────────────────────────────────
    if str(sp.mode.value if hasattr(sp.mode, 'value') else sp.mode) == "direct":
        ws3 = wb.create_sheet("Кампании")
        dir_hdrs = ["Кластер", "Тип кампании", "Заголовок", "Ключевые слова (через запятую)"]
        for ci, col in enumerate(dir_hdrs, 1):
            cell = ws3.cell(row=1, column=ci, value=col)
            cell.fill = hdr_fill
            cell.font = hdr_font

        # Group keywords by cluster
        by_cluster: dict[str, list[str]] = {}
        for kw in keywords:
            key = kw.cluster_name or "—"
            by_cluster.setdefault(key, []).append(kw.phrase)

        cluster_meta: dict[str, SemanticCluster] = {c.name: c for c in clusters}
        for ri, (cname, phrases) in enumerate(sorted(by_cluster.items()), 2):
            meta = cluster_meta.get(cname)
            row = [
                cname,
                meta.campaign_type if meta else "",
                meta.suggested_title if meta else "",
                ", ".join(phrases),
            ]
            for ci, v in enumerate(row, 1):
                ws3.cell(row=ri, column=ci, value=v)

        for ci, w in enumerate([30, 14, 38, 80], 1):
            ws3.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_csv(keywords: list) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(["Фраза", "Кластер", "Тип", "Интент", "WS", "WS_phrase", "WS_exact", "WS_order",
                     "Бренд", "Конкурент", "Сезонный", "Гео"])
    for kw in keywords:
        writer.writerow([
            kw.phrase, kw.cluster_name or "", kw.kw_type or "", kw.intent or "",
            kw.frequency_base or 0, kw.frequency_phrase or 0,
            kw.frequency_exact or 0, kw.frequency_order or 0,
            int(kw.is_branded), int(kw.is_competitor),
            int(kw.is_seasonal), int(kw.geo_dependent),
        ])
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel


def _build_txt(keywords: list) -> bytes:
    return "\n".join(kw.phrase for kw in keywords).encode("utf-8")


@router.get("/projects/{project_id}/marketing/semantic/{sem_id}/export")
def export_semantic_core(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    fmt: str = Query("xlsx", pattern="^(xlsx|csv|txt)$"),
):
    """Download semantic core as XLSX, CSV or TXT.

    XLSX has 2–3 sheets: full keyword list, cluster summary, and (Direct mode) campaign sheet.
    CSV is semicolon-delimited with BOM for direct Excel opening.
    TXT is one phrase per line (ready to paste into Wordstat or Direct).
    """
    _check_project_access(project_id, current_user, db)
    sp = _get_sem_project(sem_id, project_id, db)

    keywords, clusters = _load_export_data(sem_id, db)
    if not keywords:
        raise HTTPException(status_code=422, detail="Нет ключевых слов для экспорта")

    safe_name = _safe_filename(sp.name)

    try:
        if fmt == "xlsx":
            content = _build_xlsx(sp, keywords, clusters)
            log_event(project_id, current_user, "semantic_export", f"Экспорт СЯ ({len(keywords)} слов, XLSX)", db)
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
            )
        if fmt == "csv":
            content = _build_csv(keywords)
            log_event(project_id, current_user, "semantic_export", f"Экспорт СЯ ({len(keywords)} слов, CSV)", db)
            return Response(
                content=content,
                media_type="text/csv; charset=utf-8-sig",
                headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'},
            )
        # txt
        content = _build_txt(keywords)
        log_event(project_id, current_user, "semantic_export", f"Экспорт СЯ ({len(keywords)} слов, TXT)", db)
        return Response(
            content=content,
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.txt"'},
        )
    except Exception:
        logger.exception("Export failed for sem_id=%s fmt=%s", sem_id, fmt)
        raise


# ─── Niche Semantic Templates ────────────────────────────────────────────────

class NicheTemplateListItem(BaseModel):
    id: str
    name: str
    masks_count: int
    categories_count: int


class NicheTemplateResponse(BaseModel):
    id: str
    example_masks: list[str]
    mask_categories: list[str]
    example_keywords: list[str]
    negative_keywords_base: list[str]
    modifiers_commercial: list[str]
    modifiers_seo: list[str]


class NicheTemplateUpdate(BaseModel):
    example_masks: list[str] | None = None
    mask_categories: list[str] | None = None
    example_keywords: list[str] | None = None
    negative_keywords_base: list[str] | None = None
    modifiers_commercial: list[str] | None = None
    modifiers_seo: list[str] | None = None


_NICHE_NAMES_FALLBACK: dict[str, str] = {
    "ecommerce": "Интернет-магазин",
    "services_local": "Локальные услуги",
    "b2b_saas": "B2B / SaaS",
    "real_estate": "Недвижимость",
    "education": "Образование / Курсы",
    "medicine": "Медицина / Клиники",
    "auto": "Авто / Автосервис",
    "beauty": "Красота / Салоны",
    "construction": "Строительство домов",
}


def _load_niche_templates() -> dict:
    """Load templates from static data, merge with DB overrides."""
    from app.data.niche_semantic_templates import NICHE_SEMANTIC_TEMPLATES
    return NICHE_SEMANTIC_TEMPLATES


def _get_niche_names(db) -> dict[str, str]:
    """Get niche display names from unified brief templates (DB or defaults)."""
    from app.routers.brief_templates import _get_templates
    templates = _get_templates(db)
    names = {t["id"]: t["name"] for t in templates}
    # Merge with fallback for niches that exist in semantic but not in brief templates
    for niche_id, name in _NICHE_NAMES_FALLBACK.items():
        if niche_id not in names:
            names[niche_id] = name
    return names


@router.get("/marketing/niche-templates", response_model=list[NicheTemplateListItem])
def list_niche_templates(
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    """List available niche semantic templates."""
    templates = _load_niche_templates()
    niche_names = _get_niche_names(db)
    result = []
    for niche_id, tmpl in templates.items():
        result.append(NicheTemplateListItem(
            id=niche_id,
            name=niche_names.get(niche_id, niche_id),
            masks_count=len(tmpl.get("example_masks", [])),
            categories_count=len(tmpl.get("mask_categories", [])),
        ))
    return result


@router.get("/marketing/niche-templates/{niche_id}", response_model=NicheTemplateResponse)
def get_niche_template(niche_id: str, current_user: CurrentUser):
    """Get full niche template data."""
    templates = _load_niche_templates()
    tmpl = templates.get(niche_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail="Шаблон ниши не найден")
    return NicheTemplateResponse(
        id=niche_id,
        example_masks=tmpl.get("example_masks", []),
        mask_categories=tmpl.get("mask_categories", []),
        example_keywords=tmpl.get("example_keywords", []),
        negative_keywords_base=tmpl.get("negative_keywords_base", []),
        modifiers_commercial=tmpl.get("modifiers_commercial", []),
        modifiers_seo=tmpl.get("modifiers_seo", []),
    )


@router.put(
    "/projects/{project_id}/marketing/semantic/{sem_id}/niche-template",
    response_model=NicheTemplateResponse,
)
def save_niche_template_override(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    body: NicheTemplateUpdate,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Save per-project niche template overrides. Stored in SemanticProject.config JSON."""
    _check_project_access(project_id, current_user, db)
    sp = _get_sem_project(sem_id, project_id, db)

    config = sp.config or {}
    override = {}
    for field in ("example_masks", "mask_categories", "example_keywords",
                  "negative_keywords_base", "modifiers_commercial", "modifiers_seo"):
        val = getattr(body, field, None)
        if val is not None:
            override[field] = val
    config["niche_template_override"] = override
    sp.config = config
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(sp, "config")
    db.commit()

    return NicheTemplateResponse(
        id="custom",
        example_masks=override.get("example_masks", []),
        mask_categories=override.get("mask_categories", []),
        example_keywords=override.get("example_keywords", []),
        negative_keywords_base=override.get("negative_keywords_base", []),
        modifiers_commercial=override.get("modifiers_commercial", []),
        modifiers_seo=override.get("modifiers_seo", []),
    )


@router.get(
    "/projects/{project_id}/marketing/semantic/{sem_id}/niche-template",
    response_model=NicheTemplateResponse,
)
def get_project_niche_template(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    """Get effective niche template for this project (auto-detected + overrides)."""
    _check_project_access(project_id, current_user, db)
    sp = _get_sem_project(sem_id, project_id, db)

    from app.data.niche_semantic_templates import NICHE_SEMANTIC_TEMPLATES, detect_niche
    from app.models.brief import Brief

    # Auto-detect niche
    brief = db.scalar(select(Brief).where(Brief.project_id == project_id))
    niche_id = detect_niche(brief) if brief else None
    base = dict(NICHE_SEMANTIC_TEMPLATES.get(niche_id, {})) if niche_id else {}

    # Apply overrides from sp.config
    config = sp.config or {}
    override = config.get("niche_template_override", {})
    for field in ("example_masks", "mask_categories", "example_keywords",
                  "negative_keywords_base", "modifiers_commercial", "modifiers_seo"):
        if field in override:
            base[field] = override[field]

    return NicheTemplateResponse(
        id=niche_id or "custom",
        example_masks=base.get("example_masks", []),
        mask_categories=base.get("mask_categories", []),
        example_keywords=base.get("example_keywords", []),
        negative_keywords_base=base.get("negative_keywords_base", []),
        modifiers_commercial=base.get("modifiers_commercial", []),
        modifiers_seo=base.get("modifiers_seo", []),
    )


# ─── Competitors ──────────────────────────────────────────────────────────────

class CompetitorUrl(BaseModel):
    url: str = Field(..., min_length=1, max_length=2000)


class CompetitorListResponse(BaseModel):
    urls: list[str]


class CompetitorCrawlResult(BaseModel):
    url: str
    title: str
    h1: str
    description: str
    nav_items: list[str]
    anchors: list[str]
    frequent_terms: list[str]
    pages: list[dict]


@router.get(
    "/projects/{project_id}/marketing/semantic/{sem_id}/competitors",
    response_model=CompetitorListResponse,
)
def list_competitors(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    """Get competitor URLs from brief."""
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)

    from app.models.brief import Brief
    brief = db.scalar(select(Brief).where(Brief.project_id == project_id))
    urls = brief.competitors_urls if brief and brief.competitors_urls else []
    return CompetitorListResponse(urls=urls)


@router.post(
    "/projects/{project_id}/marketing/semantic/{sem_id}/competitors",
    response_model=CompetitorListResponse,
    status_code=201,
)
def add_competitor(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    body: CompetitorUrl,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Add a competitor URL to brief.competitors_urls."""
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)

    from app.models.brief import Brief
    brief = db.scalar(select(Brief).where(Brief.project_id == project_id))
    if not brief:
        raise HTTPException(status_code=422, detail="Бриф не заполнен")

    urls = list(brief.competitors_urls or [])
    normalized = body.url.strip().rstrip("/")
    if normalized not in [u.rstrip("/") for u in urls]:
        urls.append(body.url.strip())
        brief.competitors_urls = urls
        from sqlalchemy.orm.attributes import flag_modified
        flag_modified(brief, "competitors_urls")
        db.commit()
    return CompetitorListResponse(urls=brief.competitors_urls or [])


@router.delete(
    "/projects/{project_id}/marketing/semantic/{sem_id}/competitors",
    response_model=CompetitorListResponse,
)
def remove_competitor(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    body: CompetitorUrl,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Remove a competitor URL from brief.competitors_urls."""
    _check_project_access(project_id, current_user, db)
    _get_sem_project(sem_id, project_id, db)

    from app.models.brief import Brief
    brief = db.scalar(select(Brief).where(Brief.project_id == project_id))
    if not brief:
        raise HTTPException(status_code=422, detail="Бриф не заполнен")

    urls = list(brief.competitors_urls or [])
    normalized = body.url.strip().rstrip("/")
    urls = [u for u in urls if u.rstrip("/") != normalized]
    brief.competitors_urls = urls
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(brief, "competitors_urls")
    db.commit()
    return CompetitorListResponse(urls=urls)


@router.post(
    "/projects/{project_id}/marketing/semantic/{sem_id}/competitors/crawl",
    response_model=list[CompetitorCrawlResult],
)
def crawl_competitors(
    project_id: uuid.UUID,
    sem_id: uuid.UUID,
    current_user: CurrentUser,
    _: Annotated[object, NonViewerRequired],
    db: Annotated[Session, Depends(get_db)],
):
    """Crawl all competitor URLs and return structured extraction results."""
    _check_project_access(project_id, current_user, db)
    sp = _get_sem_project(sem_id, project_id, db)

    from app.models.brief import Brief
    brief = db.scalar(select(Brief).where(Brief.project_id == project_id))
    if not brief or not brief.competitors_urls:
        raise HTTPException(status_code=422, detail="Нет конкурентов для обхода")

    import asyncio

    import httpx
    from bs4 import BeautifulSoup

    from app.tasks.marketing import _extract_frequent_terms

    results: list[dict] = []

    async def _crawl_one(client: httpx.AsyncClient, url: str) -> dict | None:
        try:
            r = await client.get(url)
            if r.status_code != 200:
                return None
            soup = BeautifulSoup(r.text, "html.parser")
            title = (soup.title.string or "").strip() if soup.title else ""
            h1 = soup.find("h1")
            h1_text = h1.get_text(strip=True) if h1 else ""
            desc_tag = soup.find("meta", attrs={"name": "description"})
            desc = desc_tag["content"][:300] if desc_tag and desc_tag.get("content") else ""

            nav_items: list[str] = []
            for nav in soup.find_all("nav"):
                nav_items = [a.get_text(strip=True) for a in nav.find_all("a") if a.get_text(strip=True)][:30]
                if nav_items:
                    break

            base = f"{r.url.scheme}://{r.url.host}"
            anchors: list[str] = []
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if href.startswith("/") and not href.startswith("//"):
                    href = base + href
                if href.startswith(base):
                    text = a.get_text(strip=True)
                    if text and 2 < len(text) < 100:
                        anchors.append(text)
            anchors = list(dict.fromkeys(anchors))[:50]

            terms = _extract_frequent_terms(BeautifulSoup(r.text, "html.parser"))

            # Crawl internal pages (limited)
            internal_urls: set[str] = set()
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if href.startswith("/") and not href.startswith("//"):
                    href = base + href
                if href.startswith(base) and href != str(r.url) and len(internal_urls) < 10:
                    internal_urls.add(href)

            pages: list[dict] = []
            for iurl in list(internal_urls)[:10]:
                try:
                    ir = await client.get(iurl)
                    if ir.status_code != 200:
                        continue
                    isoup = BeautifulSoup(ir.text, "html.parser")
                    it = (isoup.title.string or "").strip() if isoup.title else ""
                    ih1 = isoup.find("h1")
                    ih1_text = ih1.get_text(strip=True) if ih1 else ""
                    if it or ih1_text:
                        pages.append({"url": iurl, "title": it, "h1": ih1_text})
                except Exception:
                    continue

            return {
                "url": url, "title": title, "h1": h1_text, "description": desc,
                "nav_items": nav_items, "anchors": anchors,
                "frequent_terms": terms, "pages": pages,
            }
        except Exception:
            return None

    async def _run():
        async with httpx.AsyncClient(timeout=15, follow_redirects=True, headers={"User-Agent": "SEODirectBot/1.0"}) as client:
            tasks = [_crawl_one(client, url) for url in brief.competitors_urls[:5]]
            return await asyncio.gather(*tasks)

    loop = asyncio.new_event_loop()
    try:
        raw = loop.run_until_complete(_run())
    finally:
        loop.close()

    results = [r for r in raw if r]

    # Save crawl results in sp.config for later use
    config = sp.config or {}
    config["competitor_crawl"] = results
    config["competitor_crawl_at"] = datetime.now(timezone.utc).isoformat()
    sp.config = config
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(sp, "config")
    db.commit()

    return results

